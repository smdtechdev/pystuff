import openpyxl, os
from openpyxl.styles import Font, Fill

workbook = openpyxl.load_workbook('Automation\\files\\Financial Sample.xlsx')
print(type(workbook))
sheetNames = workbook.sheetnames
print(sheetNames)
sheet = workbook['Sheet1']
print(type(sheet))

cell = sheet['B1'].value
print(cell)

print(sheet.max_row)
print(sheet.max_column)

for i in range(1, sheet.max_column+1):
    print(sheet.cell(row=1, column=i).value) # prints all the column names

#creating a new excel file
    
wb = openpyxl.Workbook()
print(wb.sheetnames)
sht = wb['Sheet']
sht2 = wb.create_sheet(title='newSheet', index=1)
print(type(sht))
print(type(sht2))

sht['A1'] = 'Cars'
sht['A2'] = 'BMW'
sht['A3'] = 'Ford'
sht['B1'] = 'Engine Type'
sht['B2'] = 'Petrol'
sht['B3'] = 'Electric'

sht['A1'].font = Font(size=12, b=True, i=True, color='00FF0000')
sht['B1'].font = Font(size=12, b=True, i=True, color='00FF0000')

sht2['A1'] = 'Motorbikes'
sht2['A2'] = 'Yamaha'
sht2['A3'] = 'Super'

sht2['A1'].font = Font(size=12, b=True, i=True, color='00FF0000')

wb.save('Automation\\files\\mysheet.xlsx')



